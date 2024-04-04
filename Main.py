import openpyxl
import sys
import datetime
import configparser

#設定ファイル読み込み
config = configparser.RawConfigParser()
config.read('settings.ini')

dt_now = datetime.datetime.now()

#設定ファイル読み込み
args = sys.argv
if(len(args)==2):
    env='DEFAULT'
else:
    env=args[2]

HEAD_LINE = int(config[env]['HEAD_LINE'])
OPTION_LINE = int(config[env]['OPTION_LINE'])
NULL_STRING = config[env]['NULL_STRING']

#エクセルファイル読み込み
wb = openpyxl.load_workbook(args[1], data_only=True)
ws_list = wb.worksheets
ws = wb.worksheets[0]
headerStr = ''
record = ''

#テーブル名(シート名)
tableName = ws.title
fileName=f'output/{tableName}_{dt_now.strftime('%Y%m%d%H%M%S')}_insert.sql'
f = open(fileName, 'a')

opList = {}
colList = []
recList = []
for col in ws.iter_cols():
    for cell in col:
        if(HEAD_LINE == cell.row):
            colList.append(str(cell.value))
            opList[str(cell.value)]=["nullcheck","string"]
        elif(OPTION_LINE == cell.row):
            currentColmun=ws.cell(HEAD_LINE,cell.column).value
            opList[currentColmun]=str(cell.value).split(',')
            
for row in ws.iter_rows():
    onerec = []
    for cell in row:
        if(HEAD_LINE != cell.row and OPTION_LINE != cell.row):
            planeRecord = str(cell.value)
            currentColmun=ws.cell(int(HEAD_LINE),cell.column).value
            if("string" in opList[currentColmun]):
                planeRecord = f'\'{str(cell.value)}\''
            if("nullcheck" in opList[currentColmun]):
                if(NULL_STRING==cell.value):
                    planeRecord = str(cell.value)
            onerec.append(planeRecord)
    if(len(onerec) != 0):
        recList.append(onerec)

for rec in recList:
    headerStr = f'INSERT INTO {tableName} ({','.join(colList)}) VALUES ('
    record = f'{headerStr}{','.join(rec)});'
    f.write(record+'\n')
